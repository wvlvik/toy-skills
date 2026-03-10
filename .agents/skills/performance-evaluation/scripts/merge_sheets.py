#!/usr/bin/env python3
"""
绩效考核表合并脚本

功能：
1. 合并多个个人绩效表格到工作簿
2. 修复标题月份错误 (x月 → 指定月份)
3. 修复考核周期错误 (年份错误、月份错误)
4. 更新汇总表月份

使用：
    python3 merge_sheets.py --input-dir "bech/3m/" --output "bech/3m/研发中心-3月份-前端组计划.xlsx" --month 3 --year 2026
    python3 merge_sheets.py --file "研发中心-3月份-前端组计划.xlsx" --month 3 --fix-only
"""

import argparse
import os
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from copy import copy
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("运行: pip3 install openpyxl --user")
    exit(1)


def get_last_day_of_month(year: int, month: int) -> int:
    """获取月份最后一天"""
    if month == 12:
        return 31
    if month in [1, 3, 5, 7, 8, 10, 12]:
        return 31
    if month in [4, 6, 9, 11]:
        return 30
    # 闰年判断
    if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
        return 29
    return 28


def generate_period(year: int, month: int) -> str:
    """生成考核周期字符串"""
    last_day = get_last_day_of_month(year, month)
    return f"{year}.{month}.1-{year}.{month}.{last_day}"


def extract_name_from_filename(filename: str) -> str:
    """从文件名中提取姓名
    
    支持的格式:
    - 周树梧.xlsx → 周树梧
    - 前端组_周树梧.xlsx → 周树梧
    - 2026年3月_周树梧.xlsx → 周树梧
    - 周树梧_绩效.xlsx → 周树梧
    - 3月_周树梧_前端.xlsx → 周树梧
    - 交易·业务线-李金涛·3月绩效考核表.xlsx → 李金涛
    - 交易·业务线·陈方杰-2603月绩效考核表.xlsx → 陈方杰
    - 刘奕君-2603月绩效.xlsx → 刘奕君
    
    Args:
        filename: 文件名（包含或不包含.xlsx后缀）
    
    Returns:
        提取的姓名，如果无法提取则返回原文件名（去掉后缀）
    """
    import re
    
    # 去掉.xlsx后缀
    name = filename.replace('.xlsx', '').replace('.XLSX', '')
    
    # 如果文件名就是纯姓名（通常2-4个中文字符），直接返回
    if re.match(r'^[一-龥]{2,4}$', name):
        return name
    
    # 排除词 - 用于识别非姓名部分（如部门、日期等）
    exclude_words = ['年', '月', '日', '组', '部', '门', '绩效', '考核', '前端', '后端', '测试', '研发', '中心', '交易', '业务线']
    
    # 尝试用多种分隔符分割：_ 、 · 、 -
    # 统一替换为下划线便于处理
    normalized = name.replace('·', '_').replace('-', '_')
    parts = normalized.split('_')
    
    # 过滤掉空部分
    parts = [p for p in parts if p]
    
    if len(parts) > 1:
        # 首先查找不包含排除词的姓名部分
        for part in parts:
            if re.match(r'^[一-龥]{2,4}$', part):
                if not any(word in part for word in exclude_words):
                    return part
        # 如果没找到，再返回第一个匹配的2-4字符中文部分
        for part in parts:
            if re.match(r'^[一-龥]{2,4}$', part):
                return part
    
    # 尝试查找连续的中文字符（2-4个）
    matches = re.findall(r'[一-龥]{2,4}', name)
    if matches:
        # 返回最可能是姓名的部分（通常不包含排除关键词）
        for match in matches:
            if not any(word in match for word in exclude_words):
                return match
        # 如果都被排除了，返回最后一个匹配项
        if matches:
            return matches[-1]
    
    # 无法提取，返回原文件名（去掉后缀）
    return name

def fix_title(ws, month: int) -> tuple[bool, str]:
    """修复标题月份

    Returns:
        (是否修改, 原值)
    """
    title = ws["A1"].value
    if not title:
        return False, ""

    title_str = str(title)

    # 修复 "x月" → 实际月份
    if "x月" in title_str:
        new_title = title_str.replace("x月", f"{month}月")
        ws["A1"] = new_title
        return True, title_str

    # 检查月份是否正确
    match = re.search(r"(\d+)月绩效考核表", title_str)
    if match:
        current_month = int(match.group(1))
        if current_month != month:
            new_title = title_str.replace(f"{current_month}月", f"{month}月")
            ws["A1"] = new_title
            return True, title_str

    return False, ""


def fix_period(ws, year: int, month: int) -> tuple[bool, str]:
    """修复考核周期

    Returns:
        (是否修改, 原值)
    """
    period = ws["J2"].value
    if not period:
        return False, ""

    period_str = str(period)
    expected_period = generate_period(year, month)

    # 检查是否需要修复
    needs_fix = False

    # 检查年份错误 (2025 → 2026)
    if str(year - 1) in period_str:
        needs_fix = True

    # 检查月份错误
    if f".{month}." not in period_str:
        needs_fix = True

    # 检查是否是预期的周期
    if period_str != expected_period:
        needs_fix = True

    if needs_fix:
        ws["J2"] = expected_period
        return True, period_str

    return False, ""


def fix_name(ws, sheet_name: str) -> tuple[bool, str]:
    """修复姓名字段，确保与工作表名称一致
    
    Returns:
        (是否修改, 原值)
    """
    # B2 单元格应该包含姓名
    current_name = ws['B2'].value
    
    if current_name != sheet_name:
        ws['B2'] = sheet_name
        return True, str(current_name) if current_name else ""
    
    return False, ""


def copy_sheet(src_ws, dst_ws):
    """复制工作表内容和格式"""
    # 复制所有单元格
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 复制列宽
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden = col_dim.hidden

    # 复制行高
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height
        dst_ws.row_dimensions[row_num].hidden = row_dim.hidden

    # 复制合并单元格
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))


def merge_individual_files(
    input_dir: str, output_path: str, month: int, year: int, dry_run: bool = False
):
    """合并个人表格到工作簿"""
    print("=== 绩效考核表合并处理 ===\n")

    # 查找个人表格文件
    input_path = Path(input_dir)
    individual_files = list(input_path.glob("*.xlsx"))
    # 排除汇总文件（仅排除包含"前端组"的文件，不排除个人绩效表）
    individual_files = [
        f for f in individual_files if "前端组" not in f.name
    ]

    if not individual_files:
        print(f"错误: 在 {input_dir} 中未找到个人表格文件")
        return

    print(f"找到 {len(individual_files)} 个个人表格:\n")
    for f in individual_files:
        name = extract_name_from_filename(f.name)
        print(f"  - {f.name} → {name}")
    # 加载或创建工作簿
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        print(f"\n加载现有工作簿: {output_path}")
    else:
        print(f"\n错误: 输出文件不存在: {output_path}")
        return

    print(f"\n1. 处理成员表格:\n")

    changes = []

    for file_path in individual_files:
        # 从文件名中提取姓名作为工作表名
        sheet_name = extract_name_from_filename(file_path.name)
        
        # 加载源文件
        src_wb = load_workbook(file_path)
        src_ws = src_wb.active

        # 删除已存在的工作表
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        # 创建新工作表
        dst_ws = wb.create_sheet(sheet_name)

        # 复制内容
        copy_sheet(src_ws, dst_ws)

        # 修复标题
        title_fixed, old_title = fix_title(dst_ws, month)
        if title_fixed:
            changes.append(
                f"  [{sheet_name}] 标题: '{old_title}' → '{dst_ws['A1'].value}'"
            )

        # 修复考核周期
        period_fixed, old_period = fix_period(dst_ws, year, month)
        if period_fixed:
            changes.append(
                f"  [{sheet_name}] 考核周期: '{old_period}' → '{dst_ws['J2'].value}'"
            )

        # 修复姓名字段
        name_fixed, old_name = fix_name(dst_ws, sheet_name)
        if name_fixed:
            changes.append(
                f"  [{sheet_name}] 姓名: '{old_name}' → '{sheet_name}'"
            )

        if not title_fixed and not period_fixed and not name_fixed:
            print(f"  [{sheet_name}] ✓ 无需修正")
    # 打印修改记录
    for change in changes:
        print(change)

    # 更新汇总表
    print(f"\n2. 更新汇总表:")
    if "汇总" in wb.sheetnames:
        summary_ws = wb["汇总"]
        header = summary_ws["A1"].value
        if header:
            header_str = str(header)
            # 更新月份
            old_month_match = re.search(r"(\d+)\s*月", header_str)
            if old_month_match:
                old_month = old_month_match.group(1)
                if int(old_month) != month:
                    # 格式化月份（保持原有空格格式）
                    new_header = re.sub(r"\d+\s*月", f"{month:02d} 月", header_str)
                    # 更新年份
                    if str(year) not in new_header:
                        new_header = re.sub(r"\d{4}年", f"{year}年", new_header)
                    summary_ws["A1"] = new_header
                    print(f"  汇总表标题: '{header_str}' → '{new_header}'")
                else:
                    print(f"  汇总表月份已是 {month} 月")

    # 保存
    print(f"\n3. 保存文件:")
    if not dry_run:
        wb.save(output_path)
        print(f"  ✓ 已保存: {output_path}")
    else:
        print(f"  [预览模式] 未保存")

    # 验证
    print(f"\n4. 验证结果:")
    wb = load_workbook(output_path)
    for file_path in individual_files:
        sheet_name = extract_name_from_filename(file_path.name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  [{sheet_name}] 标题: {ws['A1'].value} | 周期: {ws['J2'].value}")
        else:
            print(f"  [{sheet_name}] ❌ 未找到")

    print(f"\n=== 处理完成 ===")


def fix_existing_file(file_path: str, month: int, year: int, dry_run: bool = False):
    """仅修复现有工作簿中的问题"""
    print("=== 绩效考核表修复处理 ===\n")

    wb = load_workbook(file_path)
    print(f"加载文件: {file_path}")
    print(f"工作表: {wb.sheetnames}\n")

    print("1. 修复成员表格:\n")

    changes = []

    # 处理所有成员工作表（排除汇总）
    for sheet_name in wb.sheetnames:
        if sheet_name == "汇总":
            continue

        ws = wb[sheet_name]

        # 检查是否是绩效表（有标题）
        if not ws["A1"].value or "绩效考核表" not in str(ws["A1"].value):
            continue

        # 修复标题
        title_fixed, old_title = fix_title(ws, month)
        if title_fixed:
            changes.append(f"  [{sheet_name}] 标题: '{old_title}' → '{ws['A1'].value}'")

        # 修复考核周期
        period_fixed, old_period = fix_period(ws, year, month)
        if period_fixed:
            changes.append(
                f"  [{sheet_name}] 考核周期: '{old_period}' → '{ws['J2'].value}'"
            )

        # 修复姓名字段
        name_fixed, old_name = fix_name(ws, sheet_name)
        if name_fixed:
            changes.append(
                f"  [{sheet_name}] 姓名: '{old_name}' → '{sheet_name}'"
            )

        if not title_fixed and not period_fixed and not name_fixed:
            print(f"  [{sheet_name}] ✓ 无需修正")

    for change in changes:
        print(change)

    # 更新汇总表
    print(f"\n2. 更新汇总表:")
    if "汇总" in wb.sheetnames:
        summary_ws = wb["汇总"]
        header = summary_ws["A1"].value
        if header:
            header_str = str(header)
            old_month_match = re.search(r"(\d+)\s*月", header_str)
            if old_month_match:
                old_month = int(old_month_match.group(1))
                if old_month != month:
                    new_header = re.sub(r"\d+\s*月", f"{month:02d} 月", header_str)
                    if str(year) not in new_header:
                        new_header = re.sub(r"\d{4}年", f"{year}年", new_header)
                    summary_ws["A1"] = new_header
                    print(f"  汇总表标题: '{header_str}' → '{new_header}'")

    # 保存
    print(f"\n3. 保存文件:")
    if not dry_run:
        wb.save(file_path)
        print(f"  ✓ 已保存: {file_path}")
    else:
        print(f"  [预览模式] 未保存")

    print(f"\n=== 处理完成 ===")


def main():
    parser = argparse.ArgumentParser(description="绩效考核表合并与修复工具")

    # 模式选择
    mode_group = parser.add_mutually_exclusive_group(required=True)
    mode_group.add_argument("--input-dir", help="个人表格所在目录（合并模式）")
    mode_group.add_argument("--file", help="现有工作簿路径（仅修复模式）")

    # 参数
    parser.add_argument("--output", help="输出文件路径（合并模式）")
    parser.add_argument("--month", type=int, required=True, help="月份 (1-12)")
    parser.add_argument("--year", type=int, help="年份（默认当前年）")
    parser.add_argument("--dry-run", action="store_true", help="预览模式，不保存文件")

    args = parser.parse_args()

    # 默认年份
    import datetime

    year = args.year or datetime.datetime.now().year

    if args.input_dir:
        # 合并模式
        output = args.output or os.path.join(
            args.input_dir, f"研发中心-{args.month}月份-前端组计划.xlsx"
        )
        merge_individual_files(args.input_dir, output, args.month, year, args.dry_run)
    else:
        # 仅修复模式
        fix_existing_file(args.file, args.month, year, args.dry_run)


if __name__ == "__main__":
    main()
