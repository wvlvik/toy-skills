#!/usr/bin/env python3
"""
绩效Excel数据处理脚本

功能：
1. 修复标题和考核周期格式
2. 查询TAPD缺陷数据并生成URL
3. 计算绩效系数
4. 更新汇总表

用法：
    python3 process_performance.py --file "绩效表.xlsx" --month 2 --year 2026
"""

import sys
import os
import argparse
import calendar
from typing import Dict, List, Optional

import openpyxl

# 添加 tapd-idle scripts 目录到路径
script_dir = os.path.dirname(os.path.abspath(__file__))
# script_dir = .../performance-evaluation/scripts
# skills_dir = .../skills
skills_dir = os.path.dirname(os.path.dirname(script_dir))
tapd_scripts = os.path.join(skills_dir, 'tapd-idle', 'scripts')
sys.path.insert(0, tapd_scripts)

try:
    from tapd_client import TAPDClient
except ImportError:
    TAPDClient = None
    print("警告: 无法导入 TAPDClient，TAPD 查询功能将被禁用")

# 配置
WORKSPACE_ID = os.getenv('TAPD_WORKSPACE_ID', '50372234')  # 聚宝赞项目
TAPD_BASE_URL = 'https://www.tapd.cn'

# 特殊成员（上级评分在K列）
SPECIAL_COL_MEMBERS = {'陈方杰', '刘奕君'}


def calculate_coefficient(score: float) -> float:
    """计算绩效系数"""
    if score > 116:
        return 1.2
    elif score >= 106:
        return 1.1
    elif score >= 91:
        return 1.0
    elif score >= 81:
        return 0.9
    else:
        return 0.8


def get_member_sheets(wb) -> List[str]:
    """获取所有成员工作表名称（排除汇总）"""
    return [name for name in wb.sheetnames if name != '汇总']


def fix_title(ws, sheet_name: str, month: int) -> bool:
    """修复工作表标题"""
    correct_title = f"{month}月绩效考核表"
    current_title = ws['A1'].value

    if current_title != correct_title:
        ws['A1'].value = correct_title
        print(f"  [{sheet_name}] 标题: '{current_title}' -> '{correct_title}'")
        return True
    return False


def fix_period(ws, sheet_name: str, year: int, month: int) -> bool:
    """修复考核周期"""
    # 获取月份的最后一天
    last_day = calendar.monthrange(year, month)[1]
    correct_period = f"{year}.{month}.1-{year}.{month}.{last_day}"
    current_period = ws['J2'].value

    if current_period != correct_period:
        ws['J2'].value = correct_period
        print(f"  [{sheet_name}] 考核周期: '{current_period}' -> '{correct_period}'")
        return True
    return False


def fill_supervisor_rating(ws, sheet_name: str) -> bool:
    """填写缺失的上级评分（如果上级评分为空，则填入自评分）"""
    # J9: 生产质量上级评分, J10: 测试质量上级评分
    # 自评分在同一行的不同列（假设在I列）
    changes = False

    for row in [9, 10]:
        supervisor_cell = ws.cell(row=row, column=10)  # J列
        self_cell = ws.cell(row=row, column=9)  # I列（假设）

        if supervisor_cell.value is None and self_cell.value is not None:
            try:
                self_score = float(self_cell.value)
                supervisor_cell.value = self_score
                print(f"  [{sheet_name}] J{row}: 上级评分 = {self_score} (来自自评)")
                changes = True
            except (ValueError, TypeError):
                pass

    return changes


def build_tapd_url(developer: str, version_report: str, created_range: str) -> str:
    """构建TAPD缺陷查询URL

    注意：TAPD前端URL使用 queryToken 机制，需要在浏览器中手动设置过滤条件。
    这里返回的是缺陷列表基础URL。
    """
    # TAPD前端缺陷列表基础URL
    return f"{TAPD_BASE_URL}/tapd_fe/{WORKSPACE_ID}/bug/list"


def query_tapd_bugs(client, developer: str, version_report: str, created_range: str) -> Dict:
    """查询TAPD缺陷并返回数量和URL"""
    if client is None:
        return {'count': 0, 'url': build_tapd_url(developer, version_report, created_range)}

    try:
        import time
        time.sleep(0.5)  # 避免API限流

        params = {
            'workspace_id': WORKSPACE_ID,
            'de': developer,  # API参数: de = 开发人员
            'version_report': version_report,
            'created': created_range,
            'limit': 100,
            'fields': 'id,title,developer,version_report,created'
        }

        count_result = client.get_bug_count(params)
        count = count_result.get('data', {}).get('count', 0)

        return {
            'count': count,
            'url': build_tapd_url(developer, version_report, created_range)
        }
    except Exception as e:
        print(f"    查询失败: {e}")
        return {'count': 0, 'url': build_tapd_url(developer, version_report, created_range)}


def fill_tapd_urls(ws, sheet_name: str, client, created_range: str) -> bool:
    """查询TAPD缺陷并填入URL"""
    changes = False

    # 查询生产质量（线上版本）
    prod_result = query_tapd_bugs(client, sheet_name, '线上版本', created_range)
    print(f"    生产质量(线上版本): {prod_result['count']}条")

    # 查询测试质量（测试版本）
    test_result = query_tapd_bugs(client, sheet_name, '测试版本', created_range)
    print(f"    测试质量(测试版本): {test_result['count']}条")

    # 填入URL到K列
    if prod_result['url']:
        ws['K9'].value = prod_result['url']
        changes = True
    if test_result['url']:
        ws['K10'].value = test_result['url']
        changes = True

    return changes


def get_superior_scores(wb) -> Dict[str, float]:
    """从各成员工作表提取上级评分"""
    member_sheets = get_member_sheets(wb)
    results = {}

    for name in member_sheets:
        ws = wb[name]
        score_col = 11 if name in SPECIAL_COL_MEMBERS else 10

        score = None
        for row_idx in range(ws.max_row, 0, -1):
            cell_value = ws.cell(row=row_idx, column=score_col).value
            if cell_value is not None:
                try:
                    score_float = float(cell_value)
                    if score_float > 50:  # 上级评分通常 > 50
                        score = score_float
                        break
                except (ValueError, TypeError):
                    continue

        if score is not None:
            results[name] = score
            print(f"  {name}: 上级评分 = {score}")
        else:
            print(f"  {name}: 未找到上级评分")

    return results


def update_summary(wb, scores: Dict[str, float]) -> bool:
    """更新汇总表"""
    if '汇总' not in wb.sheetnames:
        print("警告: 未找到汇总工作表")
        return False

    ws = wb['汇总']
    changes = False

    # 获取汇总表中已有成员及其行号
    existing_members = {}
    for row_idx in range(3, 20):
        name = ws.cell(row=row_idx, column=4).value
        if name:
            existing_members[name] = row_idx

    # 找到最后一个有效成员行
    last_member_row = max(existing_members.values()) if existing_members else 14
    next_row = last_member_row + 1
    seq = len(existing_members)

    print("\n更新汇总表:")

    for name, score in scores.items():
        # 跳过分数为0的情况
        if score == 0:
            print(f"  {name}: 上级评分为0，保留原有分数")
            continue

        coefficient = calculate_coefficient(score)

        if name in existing_members:
            row = existing_members[name]
            old_score = ws.cell(row=row, column=5).value
            ws.cell(row=row, column=5).value = score
            ws.cell(row=row, column=6).value = coefficient
            print(f"  {name}: {old_score} -> {score}, 系数={coefficient}")
            changes = True
        else:
            # 新增记录
            seq += 1
            ws.cell(row=next_row, column=1).value = seq
            ws.cell(row=next_row, column=3).value = '前端开发'
            ws.cell(row=next_row, column=4).value = name
            ws.cell(row=next_row, column=5).value = score
            ws.cell(row=next_row, column=6).value = coefficient
            print(f"  {name}: 新增记录，分数={score}, 系数={coefficient}")
            next_row += 1
            changes = True

    return changes


def process_excel(
    filepath: str,
    month: int,
    year: int,
    fix_format: bool = True,
    query_tapd: bool = True,
    update_summary_flag: bool = True,
    dry_run: bool = False,
    output_path: Optional[str] = None
) -> bool:
    """处理绩效Excel文件"""
    print(f"=== 绩效考核表处理开始 ===\n")
    print(f"加载文件: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"错误: 无法加载文件 - {e}")
        return False

    member_sheets = get_member_sheets(wb)
    print(f"找到 {len(member_sheets)} 个成员工作表")

    # 初始化TAPD客户端
    client = None
    if query_tapd and TAPDClient:
        try:
            client = TAPDClient()
            print("TAPD客户端已连接")
        except Exception as e:
            print(f"TAPD连接失败: {e}")
            print("将跳过TAPD URL填充")

    # 计算日期范围
    last_day = calendar.monthrange(year, month)[1]
    created_range = f"{year}-{month:02d}-01~{year}-{month:02d}-{last_day}"

    changes = False

    # Phase 1: 修复格式
    if fix_format:
        print(f"\n1. 修复标题和考核周期:")
        for sheet_name in member_sheets:
            ws = wb[sheet_name]
            if fix_title(ws, sheet_name, month):
                changes = True
            if fix_period(ws, sheet_name, year, month):
                changes = True

    # Phase 2: 查询TAPD并填入URL
    if query_tapd:
        print(f"\n2. 查询TAPD缺陷数据:")
        for sheet_name in member_sheets:
            ws = wb[sheet_name]
            print(f"  [{sheet_name}]")
            if fill_tapd_urls(ws, sheet_name, client, created_range):
                changes = True

    # Phase 3: 更新汇总表
    if update_summary_flag:
        print(f"\n3. 提取上级评分:")
        scores = get_superior_scores(wb)

        if update_summary(wb, scores):
            changes = True

    # 保存文件
    if changes and not dry_run:
        output = output_path or filepath
        print(f"\n保存文件: {output}")
        wb.save(output)
        print("文件已保存")
    elif dry_run:
        print("\n[预览模式] 未保存变更")
    else:
        print("\n无需变更")

    print("\n=== 处理完成 ===")
    return True


def main():
    parser = argparse.ArgumentParser(description='绩效Excel数据处理脚本')
    parser.add_argument('--file', required=True, help='绩效表格路径')
    parser.add_argument('--month', type=int, required=True, help='月份 (1-12)')
    parser.add_argument('--year', type=int, default=2026, help='年份 (默认当前年)')
    parser.add_argument('--fix-format', action='store_true', help='仅修复格式')
    parser.add_argument('--calculate-scores', action='store_true', help='仅计算评分')
    parser.add_argument('--update-summary', action='store_true', help='仅更新汇总表')
    parser.add_argument('--dry-run', action='store_true', help='预览变更不执行')
    parser.add_argument('--output', help='输出文件路径')

    args = parser.parse_args()

    # 根据参数决定执行哪些步骤
    fix_format = True
    query_tapd = True
    update_summary_flag = True

    if args.fix_format or args.calculate_scores or args.update_summary:
        fix_format = args.fix_format
        query_tapd = args.calculate_scores or args.fix_format
        update_summary_flag = args.update_summary or args.calculate_scores

    process_excel(
        filepath=args.file,
        month=args.month,
        year=args.year,
        fix_format=fix_format,
        query_tapd=query_tapd,
        update_summary_flag=update_summary_flag,
        dry_run=args.dry_run,
        output_path=args.output
    )


if __name__ == '__main__':
    main()