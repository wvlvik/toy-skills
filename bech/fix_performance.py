#!/usr/bin/env python3
"""
绩效考核表处理脚本

执行步骤：
1. 修复标题月份错误
2. 修复考核周期错误
3. 填写缺失的上级评分
4. 查询TAPD bug数据并生成URL
"""

import sys
import os
import json
import openpyxl

# 添加 tapd-idle scripts 目录到路径
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)
tapd_scripts = os.path.join(project_root, '.agents', 'skills', 'tapd-idle', 'scripts')
sys.path.insert(0, tapd_scripts)
from tapd_client import TAPDClient

EXCEL_FILE = 'bech/2m/研发中心-2月份-前端组绩效.xlsx'
WORKSPACE_ID = '50372234'  # 聚宝赞项目
TAPD_BASE_URL = 'https://www.tapd.cn'

# 人员列表（排除汇总和刘奕君）
STAFF_LIST = [
    '吴铝', '文武', '陈方杰', '李金涛', '刘子昱',
    '戴婕妤', '熊思', '曹振辉', '方乐', '汤显文', '周树梧'
]

def fix_titles(wb):
    """修复标题月份错误"""
    fixes = {
        '吴铝': '2月绩效考核表',
        '李金涛': '2月绩效考核表',
        '方乐': '2月绩效考核表',
        '周树梧': '2月绩效考核表'
    }

    for sheet_name, correct_title in fixes.items():
        ws = wb[sheet_name]
        old_title = ws['A1'].value
        ws['A1'].value = correct_title
        print(f"  [{sheet_name}] 标题: '{old_title}' -> '{correct_title}'")

def fix_periods(wb):
    """修复考核周期错误"""
    correct_period = '2026.2.1-2026.2.28'
    sheets_to_fix = ['吴铝', '文武', '李金涛', '方乐', '汤显文']

    for sheet_name in sheets_to_fix:
        ws = wb[sheet_name]
        old_period = ws['J2'].value
        ws['J2'].value = correct_period
        print(f"  [{sheet_name}] 考核周期: '{old_period}' -> '{correct_period}'")

def fill_supervisor_ratings(wb):
    """填写缺失的上级评分"""
    # 吴铝: 生产质量上级5分, 测试质量上级5分
    ws = wb['吴铝']
    ws['J9'].value = 5
    ws['J10'].value = 5
    print(f"  [吴铝] 上级评分: J9=5, J10=5")

    # 文武: 生产质量上级10分, 测试质量上级10分
    ws = wb['文武']
    ws['J9'].value = 10
    ws['J10'].value = 10
    print(f"  [文武] 上级评分: J9=10, J10=10")

def query_tapd_bugs(client, owner, version_report, created_range):
    """查询TAPD bug并返回URL"""
    try:
        import time
        time.sleep(0.5)  # 避免API限流

        # 查询bug - TAPD使用current_owner字段存储处理人
        params = {
            'workspace_id': WORKSPACE_ID,
            'current_owner': owner,  # 使用current_owner过滤
            'version_report': version_report,
            'created': created_range,
            'limit': 100,
            'fields': 'id,title,current_owner,version_report,created'
        }

        result = client.get_bug(params)
        count_result = client.get_bug_count(params)

        count = count_result.get('data', {}).get('count', 0)

        # 构建TAPD URL
        # URL格式: https://www.tapd.cn/tapd_fe/50372234/bug/list?filter=current_owner="名字" and version_report="版本" and created="时间范围"
        url = f"{TAPD_BASE_URL}/tapd_fe/{WORKSPACE_ID}/bug/list?"

        # 添加过滤条件
        filters = []
        if owner:
            filters.append(f'current_owner="{owner}"')
        if version_report:
            filters.append(f'version_report="{version_report}"')
        if created_range:
            filters.append(f'created="{created_range}"')

        if filters:
            filter_str = ' and '.join(filters)
            url += f'filter={filter_str}'

        return {
            'count': count,
            'url': url
        }
    except Exception as e:
        print(f"    查询失败: {e}")
        return {'count': 0, 'url': ''}

def fill_tapd_urls(wb, client):
    """查询TAPD bug数据并填入URL"""
    created_range = '2026-02-01~2026-02-28'

    for staff in STAFF_LIST:
        ws = wb[staff]
        print(f"  [{staff}]")

        # 查询生产质量（线上版本）
        prod_result = query_tapd_bugs(client, staff, '线上版本', created_range)
        print(f"    生产质量(线上版本): {prod_result['count']}条")

        # 查询测试质量（测试版本）
        test_result = query_tapd_bugs(client, staff, '测试版本', created_range)
        print(f"    测试质量(测试版本): {test_result['count']}条")

        # 填入URL到K列（第11列）
        # Row 9: 生产质量 -> K9
        # Row 10: 测试质量 -> K10
        if prod_result['url']:
            ws['K9'].value = prod_result['url']
        if test_result['url']:
            ws['K10'].value = test_result['url']

def verify_changes(wb):
    """验证修改结果"""
    print("\n=== 验证结果 ===")

    # 1. 验证标题
    print("\n1. 标题验证:")
    for sheet_name in wb.sheetnames:
        if sheet_name == '汇总':
            continue
        ws = wb[sheet_name]
        title = ws['A1'].value
        status = '✓' if '2月' in str(title) else '✗'
        print(f"  {status} {sheet_name}: {title}")

    # 2. 验证考核周期
    print("\n2. 考核周期验证:")
    for sheet_name in wb.sheetnames:
        if sheet_name == '汇总':
            continue
        ws = wb[sheet_name]
        period = ws['J2'].value
        status = '✓' if period == '2026.2.1-2026.2.28' else '✗'
        print(f"  {status} {sheet_name}: {period}")

    # 3. 验证上级评分
    print("\n3. 上级评分验证:")
    for sheet_name in ['吴铝', '文武']:
        ws = wb[sheet_name]
        j9 = ws['J9'].value
        j10 = ws['J10'].value
        print(f"  {sheet_name}: J9={j9}, J10={j10}")

def main():
    print("=== 绩效考核表处理开始 ===\n")

    # 加载Excel文件
    print(f"加载文件: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Step 1: 修复标题
    print("\n1. 修复标题月份错误:")
    fix_titles(wb)

    # Step 2: 修复考核周期
    print("\n2. 修复考核周期错误:")
    fix_periods(wb)

    # Step 3: 填写上级评分
    print("\n3. 填写缺失的上级评分:")
    fill_supervisor_ratings(wb)

    # Step 4: 查询TAPD并填入URL
    print("\n4. 查询TAPD bug数据并填入URL:")
    try:
        client = TAPDClient()
        fill_tapd_urls(wb, client)
    except Exception as e:
        print(f"  TAPD连接失败: {e}")
        print("  跳过TAPD URL填充")

    # 保存文件
    print("\n保存文件...")
    wb.save(EXCEL_FILE)
    print("文件已保存")

    # 验证
    verify_changes(wb)

    print("\n=== 处理完成 ===")

if __name__ == '__main__':
    main()