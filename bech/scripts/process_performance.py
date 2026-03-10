#!/usr/bin/env python3
"""
绩效Excel数据处理脚本

从各成员工作表中提取"上级评分"合计值，更新汇总表的绩效分数和系数。
"""

import openpyxl
from typing import Dict


def calculate_coefficient(score: float) -> float:
    """计算绩效系数"""
    if score > 116:
        return 1.2
    elif score >= 106:
        return 1.1
    elif score >= 91:
        return 1.0
    elif score >= 81:
        return 0.9
    else:
        return 0.8


def get_superior_scores(wb) -> Dict[str, float]:
    """从成员工作表提取上级评分"""
    member_sheets = ['吴铝', '文武', '陈方杰', '李金涛', '刘子昱', '戴婕妤',
                     '熊思', '曹振辉', '方乐', '汤显文', '刘奕君', '周树梧',
                     '唐尧', '陈铁梁']

    special_col_members = {'陈方杰', '刘奕君'}
    results = {}

    for name in member_sheets:
        if name not in wb.sheetnames:
            print(f"警告: 工作表 '{name}' 不存在，跳过")
            continue

        ws = wb[name]
        score_col = 11 if name in special_col_members else 10

        score = None
        for row_idx in range(ws.max_row, 0, -1):
            cell_value = ws.cell(row=row_idx, column=score_col).value
            if cell_value is not None:
                try:
                    score_float = float(cell_value)
                    if score_float > 50:
                        score = score_float
                        break
                except (ValueError, TypeError):
                    continue

        if score is not None:
            results[name] = score
            print(f"  {name}: 上级评分 = {score}")
        else:
            print(f"  {name}: 未找到上级评分")

    return results


def update_summary(wb, scores: Dict[str, float]):
    """更新汇总表"""
    ws = wb['汇总']

    # 获取汇总表中已有成员及其行号
    existing_members = {}
    for row_idx in range(3, 20):
        name = ws.cell(row=row_idx, column=4).value
        if name:
            existing_members[name] = row_idx

    # 找到最后一个有效成员行
    last_member_row = max(existing_members.values()) if existing_members else 14
    next_row = last_member_row + 1
    seq = len(existing_members)

    print("\n更新汇总表:")

    for name, score in scores.items():
        if name == '吴铝' and score == 0:
            print(f"  {name}: 上级评分为0，保留原有分数")
            continue

        coefficient = calculate_coefficient(score)

        if name in existing_members:
            row = existing_members[name]
            old_score = ws.cell(row=row, column=5).value
            ws.cell(row=row, column=5).value = score
            ws.cell(row=row, column=6).value = coefficient
            print(f"  {name}: {old_score} -> {score}, 系数={coefficient}")
        else:
            # 新增记录 - 只填写非合并单元格的列
            seq += 1
            # Col 1 (序号) - 不是合并单元格
            ws.cell(row=next_row, column=1).value = seq
            # Col 2 (部门) - 是合并单元格B3:B18的一部分，不需要填写
            # Col 3 (岗位)
            ws.cell(row=next_row, column=3).value = '前端开发'
            # Col 4 (姓名)
            ws.cell(row=next_row, column=4).value = name
            # Col 5 (绩效分数)
            ws.cell(row=next_row, column=5).value = score
            # Col 6 (绩效系数)
            ws.cell(row=next_row, column=6).value = coefficient
            print(f"  {name}: 新增记录，分数={score}, 系数={coefficient}")
            next_row += 1


def main():
    filepath = 'bech/2m/研发中心-2月份-前端组绩效.xlsx'

    print(f"读取文件: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    print("\n提取上级评分:")
    scores = get_superior_scores(wb)

    update_summary(wb, scores)

    output_path = filepath.replace('.xlsx', '_updated.xlsx')
    wb.save(output_path)
    print(f"\n已保存到: {output_path}")

    print("\n=== 绩效系数规则 ===")
    print(">116: 1.2")
    print("106-115: 1.1")
    print("91-105: 1.0")
    print("81-90: 0.9")
    print("<=80: 0.8")


if __name__ == '__main__':
    main()